"""
compare_iterations.py

Compares the 'Overall Avg' column across custom20_iter1, iter2, and iter3
for avg_ files only.

Usage:
    python compare_iterations.py --base-dir /path/to/parent/folder

Output:
    iteration_comparison.xlsx with three sheets:
      - Summary  : one row per file — mean, std dev, correlation, % diff
      - Detail   : letter-by-letter values, absolute diff, % diff
      - Stats    : std dev and correlation breakdown per file pair
"""

import argparse
import os
import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ITERATIONS = ["custom20_iter1", "custom20_iter2", "custom20_iter3"]
ITER_LABELS = ["Iter1", "Iter2", "Iter3"]
PAIRS = [
    ("custom20_iter1", "custom20_iter2", "Iter2−Iter1"),
    ("custom20_iter2", "custom20_iter3", "Iter3−Iter2"),
    ("custom20_iter1", "custom20_iter3", "Iter3−Iter1"),
]


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def extract_id_and_grade(filename):
    match = re.match(r"avg_(\d+)\s*-\s*(Grade \d+)\.xlsx", filename, re.IGNORECASE)
    if match:
        return int(match.group(1)), match.group(2)
    return None, None


def load_overall_avg(filepath):
    df = pd.read_excel(filepath, usecols=["Letter", "Overall Avg"])
    df = df.dropna(subset=["Letter", "Overall Avg"])
    df["Letter"] = df["Letter"].astype(str).str.strip()
    return df.set_index("Letter")["Overall Avg"]


def build_comparison(base_dir):
    records = {}
    for iter_name in ITERATIONS:
        iter_path = os.path.join(base_dir, iter_name)
        if not os.path.isdir(iter_path):
            print(f"  WARNING: Directory not found: {iter_path}")
            continue
        for fname in os.listdir(iter_path):
            if not fname.startswith("avg_") or not fname.endswith(".xlsx"):
                continue
            file_id, grade = extract_id_and_grade(fname)
            if file_id is None:
                continue
            try:
                series = load_overall_avg(os.path.join(iter_path, fname))
            except Exception as e:
                print(f"  WARNING: Could not read {fname}: {e}")
                continue
            if file_id not in records:
                records[file_id] = {"grade": grade}
            records[file_id][iter_name] = series
    return records


# ---------------------------------------------------------------------------
# Statistics helpers
# ---------------------------------------------------------------------------

def safe_corr(s1, s2):
    """Pearson correlation between two Series aligned on their index."""
    if s1 is None or s2 is None:
        return None
    combined = pd.DataFrame({"a": s1, "b": s2}).dropna()
    if len(combined) < 2:
        return None
    return round(combined["a"].corr(combined["b"]), 4)


def pct_diff(new_val, old_val):
    """Percentage change from old_val to new_val."""
    if old_val is None or new_val is None or old_val == 0:
        return None
    return round((new_val - old_val) / abs(old_val) * 100, 2)


def letter_pct_diff(v_new, v_old):
    if v_new is None or v_old is None or v_old == 0:
        return None
    return round((v_new - v_old) / abs(v_old) * 100, 2)


def cross_iter_std(iters):
    """
    For each letter, compute std dev across the 3 iteration values.
    Return the mean of those per-letter std devs (summarises how stable
    each letter's score is across iterations).
    """
    combined = pd.DataFrame({i: s for i, s in enumerate(iters) if s is not None})
    if combined.shape[1] < 2:
        return None
    per_letter_std = combined.std(axis=1, ddof=1)
    return round(per_letter_std.mean(), 2)


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

def make_styles():
    thin = Side(style="thin", color="BFBFBF")
    return {
        "border":        Border(left=thin, right=thin, top=thin, bottom=thin),
        "header_font":   Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "normal_font":   Font(name="Arial", size=10),
        "header_fill":   PatternFill("solid", fgColor="1F4E79"),
        "section_fill":  PatternFill("solid", fgColor="2E75B6"),
        "even_fill":     PatternFill("solid", fgColor="EBF3FB"),
        "odd_fill":      PatternFill("solid", fgColor="FFFFFF"),
        "diff_pos_fill": PatternFill("solid", fgColor="C6EFCE"),
        "diff_neg_fill": PatternFill("solid", fgColor="FFC7CE"),
        "sep_fill":      PatternFill("solid", fgColor="D9D9D9"),
        "center":        Alignment(horizontal="center", vertical="center"),
        "pct_font":      Font(name="Arial", size=10, italic=True),
    }


def write_header_row(ws, headers, col_widths, s):
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = s["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 30


def style_cell(cell, val, col_idx, diff_cols, fill, s):
    cell.font = s["normal_font"]
    cell.border = s["border"]
    cell.alignment = s["center"]
    if col_idx in diff_cols and val is not None:
        cell.fill = s["diff_pos_fill"] if val > 0 else (s["diff_neg_fill"] if val < 0 else fill)
    else:
        cell.fill = fill


# ---------------------------------------------------------------------------
# Sheet 1 — Summary
# ---------------------------------------------------------------------------

def write_summary_sheet(wb, records):
    ws = wb.create_sheet(title="Summary", index=0)
    s  = make_styles()

    headers = [
        "File ID", "Grade",
        "Iter1\nMean", "Iter2\nMean", "Iter3\nMean",
        "Std Dev\n(across iters)",
        "Corr\nIter1↔2", "Corr\nIter2↔3", "Corr\nIter1↔3",
        "Δ Iter2−Iter1", "Δ Iter3−Iter2", "Δ Iter3−Iter1",
        "% Δ Iter2−Iter1", "% Δ Iter3−Iter2", "% Δ Iter3−Iter1",
    ]
    col_widths = [10, 12, 12, 12, 12, 16, 12, 12, 12, 14, 14, 14, 16, 16, 16]
    write_header_row(ws, headers, col_widths, s)

    diff_cols     = {10, 11, 12}
    pct_diff_cols = {13, 14, 15}

    for row, fid in enumerate(sorted(records.keys()), start=2):
        rec = records[fid]
        iters = [rec.get(it) for it in ITERATIONS]

        means = [round(s_.mean(), 2) if s_ is not None else None for s_ in iters]
        m1, m2, m3 = means

        # Std dev across the 3 iteration values for each letter, then averaged
        stds  = [cross_iter_std(iters)]

        corrs = [safe_corr(iters[0], iters[1]),
                 safe_corr(iters[1], iters[2]),
                 safe_corr(iters[0], iters[2])]

        abs_diffs = [
            round(m2 - m1, 2) if m2 is not None and m1 is not None else None,
            round(m3 - m2, 2) if m3 is not None and m2 is not None else None,
            round(m3 - m1, 2) if m3 is not None and m1 is not None else None,
        ]
        pct_diffs = [pct_diff(m2, m1), pct_diff(m3, m2), pct_diff(m3, m1)]

        row_vals = [fid, rec["grade"], *means, stds[0], *corrs, *abs_diffs, *pct_diffs]
        fill = s["even_fill"] if row % 2 == 0 else s["odd_fill"]

        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            if col_idx in diff_cols | pct_diff_cols and val is not None:
                cell.fill = s["diff_pos_fill"] if val > 0 else (s["diff_neg_fill"] if val < 0 else fill)
            else:
                cell.fill = fill
            cell.font   = s["normal_font"]
            cell.border = s["border"]
            cell.alignment = s["center"]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(records) + 1}"


# ---------------------------------------------------------------------------
# Sheet 2 — Detail (letter-by-letter)
# ---------------------------------------------------------------------------

def write_detail_sheet(wb, records):
    ws = wb.create_sheet(title="Detail")
    s  = make_styles()

    headers = [
        "Letter", "File ID", "Grade",
        "Iter1", "Iter2", "Iter3",
        "Std Dev\n(across iters)",
        "Δ Iter2−Iter1", "Δ Iter3−Iter2", "Δ Iter3−Iter1",
        "% Δ Iter2−Iter1", "% Δ Iter3−Iter2", "% Δ Iter3−Iter1",
    ]
    col_widths = [10, 10, 12, 12, 12, 12, 16, 14, 14, 14, 16, 16, 16]
    write_header_row(ws, headers, col_widths, s)

    diff_cols     = {8, 9, 10}
    pct_diff_cols = {11, 12, 13}

    row = 2
    for i, fid in enumerate(sorted(records.keys())):
        rec = records[fid]
        iter1 = rec.get(ITERATIONS[0])
        iter2 = rec.get(ITERATIONS[1])
        iter3 = rec.get(ITERATIONS[2])

        letters = sorted(set().union(
            iter1.index if iter1 is not None else [],
            iter2.index if iter2 is not None else [],
            iter3.index if iter3 is not None else [],
        ))

        for letter in letters:
            v1 = iter1.get(letter) if iter1 is not None else None
            v2 = iter2.get(letter) if iter2 is not None else None
            v3 = iter3.get(letter) if iter3 is not None else None

            abs_diffs = [
                round(v2 - v1, 2) if v2 is not None and v1 is not None else None,
                round(v3 - v2, 2) if v3 is not None and v2 is not None else None,
                round(v3 - v1, 2) if v3 is not None and v1 is not None else None,
            ]
            pct_diffs = [
                letter_pct_diff(v2, v1),
                letter_pct_diff(v3, v2),
                letter_pct_diff(v3, v1),
            ]

            vals = [v for v in [v1, v2, v3] if v is not None]
            letter_std = round(pd.Series(vals).std(), 2) if len(vals) >= 2 else None

            fill = s["even_fill"] if i % 2 == 0 else s["odd_fill"]
            row_vals = [letter, fid, rec["grade"], v1, v2, v3, letter_std, *abs_diffs, *pct_diffs]

            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                if col_idx in diff_cols | pct_diff_cols and val is not None:
                    cell.fill = s["diff_pos_fill"] if val > 0 else (s["diff_neg_fill"] if val < 0 else fill)
                else:
                    cell.fill = fill
                cell.font      = s["normal_font"]
                cell.border    = s["border"]
                cell.alignment = s["center"]
            row += 1

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row, column=col_idx).fill = s["sep_fill"]
        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"


# ---------------------------------------------------------------------------
# Sheet 3 — Stats (std dev + correlation deep-dive)
# ---------------------------------------------------------------------------

def write_stats_sheet(wb, records):
    ws = wb.create_sheet(title="Stats")
    s  = make_styles()

    headers = [
        "File ID", "Grade",
        "Std Dev\n(across iters)",
        "Corr Iter1↔2", "Corr Iter2↔3", "Corr Iter1↔3",
        "Iter1 Min", "Iter1 Max",
        "Iter2 Min", "Iter2 Max",
        "Iter3 Min", "Iter3 Max",
    ]
    col_widths = [10, 12, 16, 14, 14, 14, 11, 11, 11, 11, 11, 11]
    write_header_row(ws, headers, col_widths, s)

    HIGH_CORR = PatternFill("solid", fgColor="C6EFCE")
    LOW_CORR  = PatternFill("solid", fgColor="FFC7CE")
    corr_cols = {4, 5, 6}

    for row, fid in enumerate(sorted(records.keys()), start=2):
        rec   = records[fid]
        iters = [rec.get(it) for it in ITERATIONS]

        std  = cross_iter_std(iters)
        corrs = [
            safe_corr(iters[0], iters[1]),
            safe_corr(iters[1], iters[2]),
            safe_corr(iters[0], iters[2]),
        ]
        mins = [round(s_.min(), 2) if s_ is not None else None for s_ in iters]
        maxs = [round(s_.max(), 2) if s_ is not None else None for s_ in iters]

        row_vals = [fid, rec["grade"], std, *corrs,
                    mins[0], maxs[0], mins[1], maxs[1], mins[2], maxs[2]]
        fill = s["even_fill"] if row % 2 == 0 else s["odd_fill"]

        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font      = s["normal_font"]
            cell.border    = s["border"]
            cell.alignment = s["center"]
            if col_idx in corr_cols and val is not None:
                cell.fill = HIGH_CORR if val >= 0.9 else (LOW_CORR if val < 0.7 else fill)
            else:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(records) + 1}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Compare Overall Avg across iterations with std dev, correlation, and % diff.")
    parser.add_argument("--base-dir", required=True, help="Folder containing custom20_iter1/2/3")
    parser.add_argument("--out-dir",  default=None,  help="Output directory (defaults to base-dir)")
    args = parser.parse_args()

    out_dir = args.out_dir or args.base_dir
    os.makedirs(out_dir, exist_ok=True)

    print("Loading avg_ files...")
    records = build_comparison(args.base_dir)
    print(f"Found {len(records)} unique file IDs.")

    wb = Workbook()
    wb.remove(wb.active)
    write_summary_sheet(wb, records)
    write_detail_sheet(wb, records)
    write_stats_sheet(wb, records)

    out_path = os.path.join(out_dir, "iteration_comparison.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
