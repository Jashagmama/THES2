"""
compare_iterations.py

Compares the 'Overall Avg' column across custom20_iter1, iter2, and iter3
for avg_ files only.

Usage:
    python compare_iterations.py --base-dir /path/to/parent/folder

Output:
    - iteration_comparison.xlsx
        - Summary sheet: mean Overall Avg per file per iteration
        - Detail sheet: letter-by-letter breakdown
"""

import argparse
import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ITERATIONS = [
    "custom20_iter1", "custom20_iter2", "custom20_iter3",
    # "custom5_iter1", "custom5_iter2", "custom5_iter3"
    # "iter1", "iter2", "iter3"
]


def extract_id_and_grade(filename):
    match = re.match(r"avg_(\d+)\s*-\s*(Grade \d+)\.xlsx", filename, re.IGNORECASE)
    if match:
        return int(match.group(1)), match.group(2)
    return None, None


def load_overall_avg(filepath):
    df = pd.read_excel(filepath, usecols=["Letter", "Overall Avg"])
    df = df.dropna(subset=["Letter", "Overall Avg"])
    df["Letter"] = df["Letter"].astype(str).str.strip()
    return df.set_index("Letter")["Overall Avg"]


def build_comparison(base_dir):
    records = {}
    for iter_name in ITERATIONS:
        iter_path = os.path.join(base_dir, iter_name)
        if not os.path.isdir(iter_path):
            print(f"  WARNING: Directory not found: {iter_path}")
            continue
        for fname in os.listdir(iter_path):
            if not fname.startswith("avg_") or not fname.endswith(".xlsx"):
                continue
            file_id, grade = extract_id_and_grade(fname)
            if file_id is None:
                continue
            try:
                series = load_overall_avg(os.path.join(iter_path, fname))
            except Exception as e:
                print(f"  WARNING: Could not read {fname}: {e}")
                continue
            if file_id not in records:
                records[file_id] = {"grade": grade}
            records[file_id][iter_name] = series
    return records


def make_styles():
    thin = Side(style="thin", color="BFBFBF")
    return {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "header_font": Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "normal_font": Font(name="Arial", size=10),
        "header_fill": PatternFill("solid", fgColor="1F4E79"),
        "even_fill": PatternFill("solid", fgColor="EBF3FB"),
        "odd_fill": PatternFill("solid", fgColor="FFFFFF"),
        "diff_pos_fill": PatternFill("solid", fgColor="C6EFCE"),
        "diff_neg_fill": PatternFill("solid", fgColor="FFC7CE"),
        "sep_fill": PatternFill("solid", fgColor="D9D9D9"),
        "center": Alignment(horizontal="center", vertical="center"),
    }


def write_summary_sheet(wb, records):
    ws = wb.create_sheet(title="Summary", index=0)
    s = make_styles()

    headers = ["File ID", "Grade",
               "Iter1 Mean Avg", "Iter2 Mean Avg", "Iter3 Mean Avg",
               "Δ Iter2−Iter1", "Δ Iter3−Iter2", "Δ Iter3−Iter1"]
    col_widths = [10, 12, 16, 16, 16, 14, 14, 14]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = s["center"]
        cell.border = s["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for row, fid in enumerate(sorted(records.keys()), start=2):
        rec = records[fid]
        means = [round(rec[i].mean(), 2) if i in rec else None for i in ITERATIONS]
        m1, m2, m3 = means
        d21 = round(m2 - m1, 2) if m2 is not None and m1 is not None else None
        d32 = round(m3 - m2, 2) if m3 is not None and m2 is not None else None
        d31 = round(m3 - m1, 2) if m3 is not None and m1 is not None else None

        fill = s["even_fill"] if row % 2 == 0 else s["odd_fill"]
        for col_idx, val in enumerate([fid, rec["grade"], m1, m2, m3, d21, d32, d31], start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = s["normal_font"]
            cell.border = s["border"]
            cell.alignment = s["center"]
            if col_idx in (6, 7, 8) and val is not None:
                cell.fill = s["diff_pos_fill"] if val > 0 else (s["diff_neg_fill"] if val < 0 else fill)
            else:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(records) + 1}"


def write_detail_sheet(wb, records):
    ws = wb.create_sheet(title="Detail")
    s = make_styles()

    headers = ["Letter", "File ID", "Grade",
               "Iter1 Overall Avg", "Iter2 Overall Avg", "Iter3 Overall Avg",
               "Δ Iter2−Iter1", "Δ Iter3−Iter2", "Δ Iter3−Iter1"]
    col_widths = [10, 10, 12, 18, 18, 18, 14, 14, 14]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = s["center"]
        cell.border = s["border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    row = 2
    for i, fid in enumerate(sorted(records.keys())):
        rec = records[fid]
        iter1, iter2, iter3 = rec.get(ITERATIONS[0]), rec.get(ITERATIONS[1]), rec.get(ITERATIONS[2])

        letters = sorted(set().union(
            iter1.index if iter1 is not None else [],
            iter2.index if iter2 is not None else [],
            iter3.index if iter3 is not None else [],
        ))

        for letter in letters:
            v1 = iter1.get(letter) if iter1 is not None else None
            v2 = iter2.get(letter) if iter2 is not None else None
            v3 = iter3.get(letter) if iter3 is not None else None
            d21 = round(v2 - v1, 2) if v2 is not None and v1 is not None else None
            d32 = round(v3 - v2, 2) if v3 is not None and v2 is not None else None
            d31 = round(v3 - v1, 2) if v3 is not None and v1 is not None else None

            fill = s["even_fill"] if i % 2 == 0 else s["odd_fill"]
            for col_idx, val in enumerate([letter, fid, rec["grade"], v1, v2, v3, d21, d32, d31], start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = s["normal_font"]
                cell.border = s["border"]
                cell.alignment = s["center"]
                if col_idx in (7, 8, 9) and val is not None:
                    cell.fill = s["diff_pos_fill"] if val > 0 else (s["diff_neg_fill"] if val < 0 else fill)
                else:
                    cell.fill = fill
            row += 1

        # Grey separator between files
        for col_idx in range(1, 10):
            ws.cell(row=row, column=col_idx).fill = s["sep_fill"]
        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{row - 1}"


def main():
    parser = argparse.ArgumentParser(description="Compare Overall Avg across iterations (avg_ files only).")
    parser.add_argument("--base-dir", required=True, help="Folder containing custom20_iter1/2/3")
    parser.add_argument("--out-dir", default=None, help="Output directory (defaults to base-dir)")
    args = parser.parse_args()

    out_dir = args.out_dir or args.base_dir
    os.makedirs(out_dir, exist_ok=True)

    print("Loading avg_ files...")
    records = build_comparison(args.base_dir)
    print(f"Found {len(records)} unique file IDs.")

    wb = Workbook()
    wb.remove(wb.active)
    write_summary_sheet(wb, records)
    write_detail_sheet(wb, records)

    out_path = os.path.join(out_dir, "iteration_comparison.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
